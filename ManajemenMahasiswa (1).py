import streamlit as st
import pandas as pd
import sqlite3
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

DB_FILE = "siakad.db"

# 1. KONFIGURASI HALAMAN
st.set_page_config(
    page_title="SIAKAD - Sistem Informasi Academic",
    page_icon="🎓",
    layout="wide"
)

# 2. STYLE CSS KUSTOM (UI Modern & Clean)
st.markdown("""
    <style>
    @import url('https://googleapis.com');
    * { font-family: 'Inter', sans-serif; }
    
    /* Tombol Utama */
    .stButton>button {
        width: 100%;
        border-radius: 8px;
        font-weight: 500;
        transition: all 0.2s;
    }
    
    /* Desain Kartu Metrik Kustom */
    .metric-container {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
        gap: 1.5rem;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        border: 1px solid #E5E7EB;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);
        display: flex;
        align-items: center;
        gap: 1rem;
    }
    .metric-icon {
        font-size: 2.5rem;
        padding: 0.5rem;
        border-radius: 10px;
    }
    .metric-data h3 { margin: 0; font-size: 0.9rem; color: #6B7280; font-weight: 500; }
    .metric-data h1 { margin: 0; font-size: 1.8rem; color: #1F2937; font-weight: 700; }
    </style>
""", unsafe_allow_html=True)


# 3. KONEKSI & INISIALISASI DATABASE
def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS mahasiswa (
            nim TEXT PRIMARY KEY,
            nama TEXT NOT EXISTS,
            email TEXT,
            jurusan TEXT,
            semester TEXT,
            ipk TEXT,
            status TEXT
        )
    """)
    conn.commit()
    conn.close()

init_db()


# 4. CLASS & OOP LOGIC
class Mahasiswa:
    def __init__(self, nim, nama, email, jurusan, semester, ipk):
        self.__nim = nim
        self.__nama = nama
        self.__email = email
        self.__jurusan = jurusan
        self.__semester = semester
        self.__ipk = ipk

    def to_dict(self):
        return {
            "NIM": self.__nim,
            "Nama": self.__nama,
            "Email": self.__email,
            "Jurusan": self.__jurusan,
            "Semester": self.__semester,
            "IPK": self.__ipk
        }

    def status(self):
        return "Mahasiswa Aktif"


class MahasiswaBeasiswa(Mahasiswa):
    def status(self):
        return "Mahasiswa Beasiswa"


# 5. FUNGSI CRUD DATABASE
def load_data_db():
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("SELECT * FROM mahasiswa", conn)
    conn.close()
    df.columns = ["NIM", "Nama", "Email", "Jurusan", "Semester", "IPK", "Status"]
    return df

def insert_data_db(data_dict, status_str):
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO mahasiswa (nim, nama, email, jurusan, semester, ipk, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (data_dict["NIM"], data_dict["Nama"], data_dict["Email"], data_dict["Jurusan"], data_dict["Semester"], data_dict["IPK"], status_str))
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        st.error(f"❌ Gagal: NIM {data_dict['NIM']} sudah digunakan oleh mahasiswa lain.")
        return False

def update_data_db(data_dict, status_str):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE mahasiswa 
        SET nama=?, email=?, jurusan=?, semester=?, ipk=?, status=?
        WHERE nim=?
    """, (data_dict["Nama"], data_dict["Email"], data_dict["Jurusan"], data_dict["Semester"], data_dict["IPK"], status_str, data_dict["NIM"]))
    conn.commit()
    conn.close()

def delete_data_db(nim):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM mahasiswa WHERE nim=?", (nim,))
    conn.commit()
    conn.close()


# 6. FUNGSI GENERATOR LAPORAN (EXCEL & PDF)
def export_to_excel(df):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Mahasiswa"
    
    # Header Styling
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Data Rows
    for r_idx, row in enumerate(df.values, 2):
        ws.append(list(row))
        
    # Auto-fit column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output)
    return output.getvalue()


def export_to_pdf(df):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PDFTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=15,
        textColor=colors.HexColor('#0F4C81')
    )
    desc_style = ParagraphStyle(
        'PDFDesc',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=25,
        textColor=colors.HexColor('#555555')
    )
    
    # Kop Dokumen
    story.append(Paragraph("LAPORAN DATA MAHASISWA AKTIF", title_style))
    story.append(Paragraph("Sistem Informasi Akademik (SIAKAD) - Laporan Internal Real-time", desc_style))
    
    # Siapkan Data Tabel PDF (Hanya kolom-kolom penting agar muat halaman)
    pdf_df = df[["NIM", "Nama", "Jurusan", "Semester", "IPK", "Status"]]
    data = [pdf_df.columns.tolist()] + pdf_df.values.tolist()
    
    # Pengaturan Ukuran Kolom Tabel (Total Lebar ~550)
    col_widths = [75, 130, 110, 55, 45, 110]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F4C81')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (4, -1), 'CENTER'), # Center untuk Semester & IPK
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F9FAFB'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
    ]))
    
    story.append(table)
    doc.build(story)
    return output.getvalue()


# VALIDASI FORM & UTILITAS
def valid_nim(nim): return bool(re.match(r"^[0-9]{5,15}$", nim))
def valid_nama(nama): return bool(re.match(r"^[A-Za-z\s]+$", nama))
def valid_email(email): return bool(re.match(r"^[\w\.-]+@[\w\.-]+\.\w+$", email))

def predikat_ipk(ipk):
    try:
        val = float(ipk)
        if val >= 3.75: return "🥇 Cumlaude"
        if val >= 3.50: return "🟢 Sangat Memuaskan"
        if val >= 3.00: return "🟡 Memuaskan"
        if val >= 2.00: return "🟠 Cukup"
        return "🔴 Kurang"
    except:
        return "-"


# ALGORITMA SEARCH & SORT
def linear_search(df, keyword):
    mask = df["Nama"].str.lower().str.contains(keyword.lower()) | df["NIM"].str.contains(keyword)
    return df[mask]

def binary_search(df, nim):
    df_sorted = df.sort_values("NIM").reset_index(drop=True)
    kiri, kanan = 0, len(df_sorted) - 1
    while kiri <= kanan:
        tengah = (kiri + kanan) // 2
        if df_sorted.loc[tengah, "NIM"] == nim:
            return pd.DataFrame([df_sorted.loc[tengah]])
        elif df_sorted.loc[tengah, "NIM"] < nim:
            kiri = tengah + 1
        else:
            kanan = tengah - 1
    return pd.DataFrame()

def merge_sort(df, kolom):
    data = df.to_dict("records")
    def ms(lst):
        if len(lst) <= 1: return lst
        mid = len(lst) // 2
        left = ms(lst[:mid])
        right = ms(lst[mid:])
        
        res, i, j = [], 0, 0
        while i < len(left) and j < len(right):
            if str(left[i][kolom]) <= str(right[j][kolom]):
                res.append(left[i]); i += 1
            else:
                res.append(right[j]); j += 1
        res.extend(left[i:]); res.extend(right[j:])
        return res
    return pd.DataFrame(ms(data))


# 7. ANTARMUKA SIDEBAR (MENU UTAMA)
with st.sidebar:
    st.markdown("<h2 style='margin-bottom:0;'>🎓 SIAKAD APP</h2>", unsafe_allow_html=True)
    st.markdown("<p style='color:gray; font-size:0.85rem;'>Sistem Informasi Akademik v4.0</p>", unsafe_allow_html=True)
    st.divider()
    
    menu = st.radio(
        "Navigasi Menu",
        ["Dashboard Utama", "Tambah Data", "Kelola Data (Edit/Hapus)", "Pencarian & Urutan"],
        index=0
    )
    st.sidebar.caption("Database: SQLite | Fitur: Ekspor Dokumen")

# Ambil data terbaru dari database
df_mhs = load_data_db()


# 8. KONTEN HALAMAN UTAMA
if menu == "Dashboard Utama":
